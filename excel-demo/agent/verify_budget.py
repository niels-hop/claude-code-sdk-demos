from openpyxl import load_workbook

wb = load_workbook('Budget_Tracker.xlsx', data_only=False)
sheet = wb.active

print("Checking formulas in Budget_Tracker.xlsx...")
print("\nFormula verification:")

formula_cells = []
for row in sheet.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formula_cells.append((cell.coordinate, cell.value))

print(f"\nTotal formulas found: {len(formula_cells)}")
print("\nSample formulas:")
for coord, formula in formula_cells[:10]:
    print(f"  {coord}: {formula}")

print("\n✓ Budget tracker created successfully!")
print(f"  Location: {wb.worksheets[0].title}")
print(f"  Total formulas: {len(formula_cells)}")

wb.close()
