#!/usr/bin/env python3
"""
Custom Excel Formula Recalculation Script for macOS
"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path

def setup_libreoffice_macro():
    """Setup LibreOffice macro for recalculation if not already configured"""
    macro_dir = os.path.expanduser('~/Library/Application Support/LibreOffice/4/user/basic/Standard')
    macro_file = os.path.join(macro_dir, 'Module1.xba')
    
    if os.path.exists(macro_file):
        try:
            with open(macro_file, 'r') as f:
                if 'RecalculateAndSave' in f.read():
                    return True
        except:
            pass
    
    os.makedirs(macro_dir, exist_ok=True)
    
    # First run LibreOffice to initialize config if needed
    try:
        subprocess.run(['/Applications/LibreOffice.app/Contents/MacOS/soffice', '--headless', '--terminate_after_init'], 
                      capture_output=True, timeout=15)
    except:
        pass
    
    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''
    
    try:
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        print(f"Macro created at: {macro_file}")
        return True
    except Exception as e:
        print(f"Failed to create macro: {e}")
        return False

def recalc_simple(filename):
    """Simple recalculation approach using LibreOffice calc mode"""
    abs_path = str(Path(filename).absolute())
    
    cmd = [
        '/Applications/LibreOffice.app/Contents/MacOS/soffice',
        '--headless',
        '--calc',
        '--convert-to', 'xlsx',
        '--outdir', str(Path(filename).parent),
        abs_path
    ]
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        print(f"LibreOffice conversion result: {result.returncode}")
        if result.stdout:
            print(f"Output: {result.stdout}")
        if result.stderr:
            print(f"Errors: {result.stderr}")
        return result.returncode == 0
    except Exception as e:
        print(f"Error running LibreOffice: {e}")
        return False

def check_excel_errors(filename):
    """Check for Excel formula errors in the file"""
    try:
        sys.path.insert(0, '/Users/thariq/code/excel-demo/venv/lib/python3.13/site-packages')
        from openpyxl import load_workbook
        
        wb = load_workbook(filename, data_only=True)
        
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break
        
        # Count formulas
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        
        wb.close()
        wb_formulas.close()
        
        result = {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'total_formulas': formula_count,
            'error_summary': {}
        }
        
        for err_type, locations in error_details.items():
            if locations:
                result['error_summary'][err_type] = {
                    'count': len(locations),
                    'locations': locations[:20]
                }
        
        return result
        
    except Exception as e:
        return {'error': str(e)}

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python recalc_custom.py <excel_file>")
        sys.exit(1)
    
    filename = sys.argv[1]
    
    if not Path(filename).exists():
        result = {'error': f'File {filename} does not exist'}
    else:
        # Check current state
        result = check_excel_errors(filename)
        
    print(json.dumps(result, indent=2))