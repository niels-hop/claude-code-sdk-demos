from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
sheet = wb.active
sheet.title = "Monthly Budget"

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
category_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
total_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
input_font = Font(color='0000FF')

sheet.column_dimensions['A'].width = 25
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15

sheet['A1'] = 'MONTHLY BUDGET TRACKER'
sheet['A1'].font = Font(bold=True, size=14)
sheet.merge_cells('A1:E1')
sheet['A1'].alignment = Alignment(horizontal='center')

sheet['A3'] = 'Category'
sheet['B3'] = 'Budgeted'
sheet['C3'] = 'Actual'
sheet['D3'] = 'Difference'
sheet['E3'] = 'Status'

for col in ['A', 'B', 'C', 'D', 'E']:
    cell = sheet[f'{col}3']
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

sheet['A5'] = 'INCOME'
sheet['A5'].font = Font(bold=True, size=11)
sheet['A5'].fill = category_fill

income_categories = ['Salary', 'Freelance/Side Income', 'Investment Income', 'Other Income']
row = 6
for category in income_categories:
    sheet[f'A{row}'] = category
    sheet[f'B{row}'] = 0
    sheet[f'B{row}'].font = input_font
    sheet[f'C{row}'] = 0
    sheet[f'C{row}'].font = input_font
    sheet[f'D{row}'] = f'=C{row}-B{row}'
    sheet[f'E{row}'] = f'=IF(D{row}>=0,"Over","Under")'
    for col in ['A', 'B', 'C', 'D', 'E']:
        sheet[f'{col}{row}'].border = thin_border
    row += 1

sheet[f'A{row}'] = 'Total Income'
sheet[f'A{row}'].font = Font(bold=True)
sheet[f'A{row}'].fill = total_fill
sheet[f'B{row}'] = f'=SUM(B6:B{row-1})'
sheet[f'C{row}'] = f'=SUM(C6:C{row-1})'
sheet[f'D{row}'] = f'=C{row}-B{row}'
sheet[f'E{row}'] = ''
for col in ['A', 'B', 'C', 'D', 'E']:
    cell = sheet[f'{col}{row}']
    cell.border = thin_border
    if col in ['B', 'C', 'D']:
        cell.font = Font(bold=True)
total_income_row = row
row += 2

sheet[f'A{row}'] = 'EXPENSES'
sheet[f'A{row}'].font = Font(bold=True, size=11)
sheet[f'A{row}'].fill = category_fill
row += 1

expense_categories = [
    'Housing (Rent/Mortgage)',
    'Utilities',
    'Groceries',
    'Transportation',
    'Insurance',
    'Healthcare',
    'Debt Payments',
    'Entertainment',
    'Dining Out',
    'Shopping',
    'Savings',
    'Emergency Fund',
    'Other Expenses'
]

expense_start_row = row
for category in expense_categories:
    sheet[f'A{row}'] = category
    sheet[f'B{row}'] = 0
    sheet[f'B{row}'].font = input_font
    sheet[f'C{row}'] = 0
    sheet[f'C{row}'].font = input_font
    sheet[f'D{row}'] = f'=C{row}-B{row}'
    sheet[f'E{row}'] = f'=IF(D{row}<=0,"Under","Over")'
    for col in ['A', 'B', 'C', 'D', 'E']:
        sheet[f'{col}{row}'].border = thin_border
    row += 1

sheet[f'A{row}'] = 'Total Expenses'
sheet[f'A{row}'].font = Font(bold=True)
sheet[f'A{row}'].fill = total_fill
sheet[f'B{row}'] = f'=SUM(B{expense_start_row}:B{row-1})'
sheet[f'C{row}'] = f'=SUM(C{expense_start_row}:C{row-1})'
sheet[f'D{row}'] = f'=C{row}-B{row}'
sheet[f'E{row}'] = ''
for col in ['A', 'B', 'C', 'D', 'E']:
    cell = sheet[f'{col}{row}']
    cell.border = thin_border
    if col in ['B', 'C', 'D']:
        cell.font = Font(bold=True)
total_expense_row = row
row += 2

sheet[f'A{row}'] = 'NET INCOME (Surplus/Deficit)'
sheet[f'A{row}'].font = Font(bold=True, size=11)
sheet[f'A{row}'].fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
sheet[f'B{row}'] = f'=B{total_income_row}-B{total_expense_row}'
sheet[f'C{row}'] = f'=C{total_income_row}-C{total_expense_row}'
sheet[f'D{row}'] = f'=C{row}-B{row}'
sheet[f'E{row}'] = f'=IF(C{row}>=0,"Surplus","Deficit")'
for col in ['A', 'B', 'C', 'D', 'E']:
    cell = sheet[f'{col}{row}']
    cell.border = thin_border
    cell.font = Font(bold=True)
row += 2

sheet[f'A{row}'] = 'Budget Summary'
sheet[f'A{row}'].font = Font(bold=True, size=11, underline='single')
row += 1
sheet[f'A{row}'] = 'Savings Rate:'
sheet[f'B{row}'] = f'=IF(C{total_income_row}>0,C{expense_start_row+10}/C{total_income_row},0)'
sheet[f'B{row}'].number_format = '0.0%'
row += 1
sheet[f'A{row}'] = 'Expense Ratio:'
sheet[f'B{row}'] = f'=IF(C{total_income_row}>0,C{total_expense_row}/C{total_income_row},0)'
sheet[f'B{row}'].number_format = '0.0%'

for row_num in range(6, sheet.max_row + 1):
    for col in ['B', 'C', 'D']:
        cell = sheet[f'{col}{row_num}']
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            cell.number_format = '$#,##0;($#,##0);-'

wb.save('Budget_Tracker.xlsx')
